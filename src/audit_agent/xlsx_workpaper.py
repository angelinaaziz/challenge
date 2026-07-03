"""Excel workpaper writer.

Bead markets "native Excel working papers" — this module produces one from the
same SampleAssessment data that drives the HTML report. Sheets:
  - Cover              headline verdict, exec summary, key findings, actions, sign-off
  - Attribute Verdicts one row per attribute with rationale + evidence coord
  - Evidence Citations flattened citations for easy pivoting
  - Reperformance      raw finding list (UAR)
  - Evidence Inventory what was ingested and who cited it
  - Decision Log       every LLM call (audit trail)
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from audit_agent.schemas import SampleAssessment, rollup_verdicts


# Bead-themed colour palette (from usebead.ai) — reused in cell fills.
PAPER = "FAF8F3"
INK = "0B362C"
ACCENT = "148F72"
ACCENT_SOFT = "E7F3EE"
PASS = "148F72"
PASS_SOFT = "E7F3EE"
FAIL = "B83A2E"
FAIL_SOFT = "F4E1DF"
WARN = "A17321"
WARN_SOFT = "F5EDD9"
BORDER = "D1CBBA"

_VERDICT_FILL = {
    "SUCCESS": PASS_SOFT,
    "FAIL": FAIL_SOFT,
    "FURTHER_EVIDENCE_REQUIRED": WARN_SOFT,
    "CONTROL_PASS": PASS_SOFT,
    "CONTROL_FAIL": FAIL_SOFT,
    "CONTROL_INCONCLUSIVE": WARN_SOFT,
}
_VERDICT_FONT = {
    "SUCCESS": PASS,
    "FAIL": FAIL,
    "FURTHER_EVIDENCE_REQUIRED": WARN,
    "CONTROL_PASS": PASS,
    "CONTROL_FAIL": FAIL,
    "CONTROL_INCONCLUSIVE": WARN,
}

_thin_border = Border(
    left=Side(style="thin", color=BORDER),
    right=Side(style="thin", color=BORDER),
    top=Side(style="thin", color=BORDER),
    bottom=Side(style="thin", color=BORDER),
)


def _header_cell(cell) -> None:
    cell.font = Font(name="SF Pro Text", size=10, bold=True, color=INK)
    cell.fill = PatternFill("solid", fgColor=ACCENT_SOFT)
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = _thin_border


def _body_cell(cell) -> None:
    cell.font = Font(name="SF Pro Text", size=10, color=INK)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _thin_border


def _set_col_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# --- Sheet builders ---------------------------------------------------------

def _build_cover(ws, sa: SampleAssessment) -> None:
    ws.title = "Cover"
    verdict = sa.control_conclusion.value

    ws["A1"] = "Bead audit working paper"
    ws["A1"].font = Font(name="SF Pro Display", size=8, bold=True, color=ACCENT)
    ws.merge_cells("A1:D1")

    ws["A2"] = sa.control
    ws["A2"].font = Font(name="SF Pro Display", size=22, bold=True, color=INK)
    ws.merge_cells("A2:D2")

    ws["A4"] = "Verdict"
    ws["A4"].font = Font(name="SF Pro Text", size=9, bold=True, color=INK)
    ws["B4"] = verdict.replace("_", " ")
    ws["B4"].font = Font(name="SF Pro Text", size=16, bold=True, color=_VERDICT_FONT[verdict])
    ws["B4"].fill = PatternFill("solid", fgColor=_VERDICT_FILL[verdict])
    ws.merge_cells("B4:D4")

    ws["A6"] = "Sample"
    ws["B6"] = sa.sample_id
    ws["A7"] = "Model"
    ws["B7"] = sa.model
    ws["A8"] = "Generated"
    ws["B8"] = sa.generated_at.strftime("%d %b %Y, %H:%M UTC")
    for r in range(6, 9):
        ws[f"A{r}"].font = Font(name="SF Pro Text", size=9, bold=True, color="6F6A5F")
        ws[f"B{r}"].font = Font(name="SF Pro Text", size=10, color=INK)

    row = 10
    if sa.executive_summary:
        ws[f"A{row}"] = "Executive summary"
        ws[f"A{row}"].font = Font(name="SF Pro Text", size=9, bold=True, color=ACCENT)
        row += 1
        ws[f"A{row}"] = sa.executive_summary
        ws[f"A{row}"].alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 60
        row += 2

    if sa.key_findings:
        ws[f"A{row}"] = "Key findings"
        ws[f"A{row}"].font = Font(name="SF Pro Text", size=9, bold=True, color=ACCENT)
        row += 1
        for f in sa.key_findings:
            ws[f"A{row}"] = f.severity.upper()
            ws[f"A{row}"].font = Font(
                name="SF Pro Text", size=9, bold=True,
                color={"pass": PASS, "warn": WARN, "fail": FAIL}[f.severity],
            )
            ws[f"B{row}"] = f.text
            ws[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1
        row += 1

    if sa.recommended_actions:
        ws[f"A{row}"] = "Recommended actions"
        ws[f"A{row}"].font = Font(name="SF Pro Text", size=9, bold=True, color=ACCENT)
        row += 1
        for i, act in enumerate(sa.recommended_actions, 1):
            ws[f"A{row}"] = str(i)
            ws[f"A{row}"].font = Font(name="SF Pro Text", size=10, bold=True, color=ACCENT)
            ws[f"B{row}"] = act
            ws[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1

    ws[f"A{row}"] = "Reviewer sign-off"
    ws[f"A{row}"].font = Font(name="SF Pro Text", size=9, bold=True, color=ACCENT)
    row += 1
    for label in ("Decision (Accept / Rework / Reject)", "Reviewer name", "Reviewer role", "Date", "Comments"):
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name="SF Pro Text", size=9, color="6F6A5F")
        ws[f"B{row}"] = ""
        ws[f"B{row}"].fill = PatternFill("solid", fgColor=PAPER)
        ws[f"B{row}"].border = _thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    _set_col_widths(ws, {1: 26, 2: 40, 3: 22, 4: 22})


def _build_attributes_sheet(wb: Workbook, sa: SampleAssessment) -> None:
    ws = wb.create_sheet("Attribute Verdicts")
    headers = ["Attribute", "Verdict", "Confidence", "Rationale", "Citations", "Exceptions considered"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        _header_cell(cell)

    for r, a in enumerate(sa.attributes, start=2):
        ws.cell(row=r, column=1, value=a.attribute_text)
        ws.cell(row=r, column=2, value=a.verdict.value)
        ws.cell(row=r, column=3, value=round(a.confidence, 2))
        ws.cell(row=r, column=4, value=a.rationale)
        ws.cell(row=r, column=5, value=len(a.evidence_refs))
        ws.cell(row=r, column=6, value="\n".join(a.exceptions_considered) if a.exceptions_considered else "")
        for c in range(1, 7):
            _body_cell(ws.cell(row=r, column=c))
        # Colour the verdict cell.
        v_cell = ws.cell(row=r, column=2)
        v_cell.fill = PatternFill("solid", fgColor=_VERDICT_FILL[a.verdict.value])
        v_cell.font = Font(name="SF Pro Text", size=10, bold=True, color=_VERDICT_FONT[a.verdict.value])
        ws.row_dimensions[r].height = 80

    _set_col_widths(ws, {1: 50, 2: 20, 3: 12, 4: 70, 5: 12, 6: 40})
    ws.freeze_panes = "A2"


def _build_citations_sheet(wb: Workbook, sa: SampleAssessment) -> None:
    ws = wb.create_sheet("Evidence Citations")
    headers = ["Attribute", "File", "Locator", "Observation"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=i, value=h))

    r = 2
    for a in sa.attributes:
        for e in a.evidence_refs:
            ws.cell(row=r, column=1, value=a.attribute_text)
            ws.cell(row=r, column=2, value=e.file)
            ws.cell(row=r, column=3, value=e.locator)
            ws.cell(row=r, column=4, value=e.observation)
            for c in range(1, 5):
                _body_cell(ws.cell(row=r, column=c))
            r += 1

    _set_col_widths(ws, {1: 45, 2: 40, 3: 32, 4: 70})
    ws.freeze_panes = "A2"


def _build_reperformance_sheet(wb: Workbook, sa: SampleAssessment) -> None:
    if not sa.reperformance_notes:
        return
    ws = wb.create_sheet("Reperformance")
    ws["A1"] = "Deterministic reperformance"
    ws["A1"].font = Font(name="SF Pro Text", size=12, bold=True, color=ACCENT)
    ws["A2"] = "The deterministic re-check output. This is the ground-truth layer the LLM judge sees."
    ws["A2"].font = Font(name="SF Pro Text", size=9, color="6F6A5F")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:D2")
    ws["A4"] = sa.reperformance_notes
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A4:D4")
    ws.row_dimensions[4].height = 100
    _set_col_widths(ws, {1: 30, 2: 30, 3: 30, 4: 30})


def _build_inventory_sheet(wb: Workbook, sa: SampleAssessment) -> None:
    if not sa.evidence_inventory:
        return
    ws = wb.create_sheet("Evidence Inventory")
    headers = ["File", "Type", "Size (bytes)", "Extraction summary", "Cited by attributes"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=i, value=h))
    for r, ei in enumerate(sa.evidence_inventory, start=2):
        ws.cell(row=r, column=1, value=ei.file)
        ws.cell(row=r, column=2, value=ei.kind)
        ws.cell(row=r, column=3, value=ei.bytes_size)
        ws.cell(row=r, column=4, value=ei.extraction_summary)
        ws.cell(row=r, column=5, value=", ".join(ei.cited_by_attributes) or "— uncited —")
        for c in range(1, 6):
            _body_cell(ws.cell(row=r, column=c))
        if not ei.cited_by_attributes:
            ws.cell(row=r, column=5).font = Font(name="SF Pro Text", size=10, color=WARN, italic=True)
    _set_col_widths(ws, {1: 55, 2: 14, 3: 14, 4: 55, 5: 40})
    ws.freeze_panes = "A2"


def _build_decision_log_sheet(wb: Workbook, trace_path: Path) -> None:
    if not trace_path.exists():
        return
    rows: list[dict] = []
    for line in trace_path.read_text().splitlines():
        if line.strip():
            rows.append(json.loads(line))
    if not rows:
        return
    ws = wb.create_sheet("Decision Log")
    headers = ["#", "Timestamp", "Purpose", "Model", "Input tokens", "Cache read", "Output tokens", "Cost USD", "Latency (ms)"]
    for i, h in enumerate(headers, 1):
        _header_cell(ws.cell(row=1, column=i, value=h))
    for i, r in enumerate(rows, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=r.get("ts", ""))
        ws.cell(row=i + 1, column=3, value=r.get("purpose", ""))
        ws.cell(row=i + 1, column=4, value=r.get("model", ""))
        ws.cell(row=i + 1, column=5, value=r.get("input_tokens") or 0)
        ws.cell(row=i + 1, column=6, value=r.get("cache_read_tokens") or 0)
        ws.cell(row=i + 1, column=7, value=r.get("output_tokens") or 0)
        cost_cell = ws.cell(row=i + 1, column=8, value=r.get("cost_usd") or 0)
        cost_cell.number_format = '"$"#,##0.0000'
        ws.cell(row=i + 1, column=9, value=r.get("latency_ms") or 0)
        for c in range(1, 10):
            _body_cell(ws.cell(row=i + 1, column=c))
    # Total row
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True, color=INK)
    for c in (5, 6, 7, 8, 9):
        col_letter = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        cell.font = Font(bold=True, color=INK)
        if c == 8:
            cell.number_format = '"$"#,##0.0000'
    _set_col_widths(ws, {1: 6, 2: 30, 3: 55, 4: 24, 5: 14, 6: 14, 7: 14, 8: 12, 9: 14})
    ws.freeze_panes = "A2"


# --- Public API -------------------------------------------------------------

def build_workpaper(sa: SampleAssessment, out_path: Path, trace_path: Path | None = None) -> Path:
    """Write an Excel workpaper for a single sample. Returns the path written."""
    wb = Workbook()
    default = wb.active
    _build_cover(default, sa)
    _build_attributes_sheet(wb, sa)
    _build_citations_sheet(wb, sa)
    _build_reperformance_sheet(wb, sa)
    _build_inventory_sheet(wb, sa)
    if trace_path is not None:
        _build_decision_log_sheet(wb, trace_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_workpapers_for_run(run_dir: Path) -> list[Path]:
    """Write one workpaper.xlsx per sample under run_dir."""
    trace = run_dir / "trace.jsonl"
    out_paths: list[Path] = []
    for sub in sorted(run_dir.iterdir()):
        if not sub.is_dir():
            continue
        ap = sub / "assessment.json"
        if not ap.exists():
            continue
        raw = json.loads(ap.read_text())
        if "control_conclusion" not in raw:
            from audit_agent.schemas import Verdict
            raw["control_conclusion"] = rollup_verdicts(
                [Verdict(a["verdict"]) for a in raw.get("attributes", [])]
            ).value
        sa = SampleAssessment(**raw)
        out = sub / "workpaper.xlsx"
        build_workpaper(sa, out, trace_path=trace if trace.exists() else None)
        out_paths.append(out)
    return out_paths
