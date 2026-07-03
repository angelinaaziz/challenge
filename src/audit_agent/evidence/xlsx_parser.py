"""Deterministic xlsx parsing with cell-coordinate preservation.

We do NOT send the whole xlsx to the LLM. Sheets get parsed into typed Python
structures with cell addresses attached, so downstream code can join them and
produce citation strings like `Access Review!E13`."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from audit_agent.schemas import XlsxFacts, XlsxSheetSummary


@dataclass
class XlsxRow:
    """One row from a sheet, keyed by header name, with per-cell coordinates."""
    row_index: int  # 1-based openpyxl row number
    values: dict[str, Any]  # header → cell value (typed)
    coords: dict[str, str]  # header → openpyxl coordinate (e.g. "E13")


@dataclass
class XlsxSheet:
    name: str
    headers: list[str]
    header_row: int
    rows: list[XlsxRow] = field(default_factory=list)
    # Cover-page-style key/value pairs (when the sheet is A: label / B: value)
    kv: dict[str, tuple[Any, str]] = field(default_factory=dict)

    @property
    def is_tabular(self) -> bool:
        return bool(self.headers and self.rows)


@dataclass
class XlsxWorkbook:
    path: Path
    filename: str
    sheets: dict[str, XlsxSheet]

    def sheet(self, name: str) -> XlsxSheet | None:
        return self.sheets.get(name)


# --- Parsing -----------------------------------------------------------------

def _find_header_row(ws) -> int | None:
    """Best-effort: find a row that looks like column headers.

    Heuristic: first row where >= 3 consecutive cells starting from A are non-empty strings.
    Handles files where row 1 is a title and row 2 is blank (like these workbooks).
    """
    for r in range(1, min(ws.max_row + 1, 15)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        strs = [v for v in row_vals if isinstance(v, str) and v.strip()]
        if len(strs) >= 3 and all(isinstance(v, str) for v in row_vals[: len(strs)]):
            return r
    return None


def _parse_tabular_sheet(ws, header_row: int) -> tuple[list[str], list[XlsxRow]]:
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            break
        headers.append(str(v))

    rows: list[XlsxRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        values: dict[str, Any] = {}
        coords: dict[str, str] = {}
        any_val = False
        for i, h in enumerate(headers):
            cell = ws.cell(row=r, column=i + 1)
            v = cell.value
            if v is not None:
                any_val = True
            values[h] = v
            coords[h] = cell.coordinate
        if any_val:
            rows.append(XlsxRow(row_index=r, values=values, coords=coords))
    return headers, rows


def _parse_kv_sheet(ws) -> dict[str, tuple[Any, str]]:
    """Extract A/B key-value pairs from a cover-page-style sheet."""
    kv: dict[str, tuple[Any, str]] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        if isinstance(label, str) and label.strip() and value is not None:
            kv[label.strip().rstrip(":")] = (value, ws.cell(row=r, column=2).coordinate)
    return kv


def load_xlsx(path: Path) -> XlsxWorkbook:
    wb = load_workbook(path, data_only=True, read_only=False)
    sheets: dict[str, XlsxSheet] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = _find_header_row(ws)
        if header_row is not None:
            headers, rows = _parse_tabular_sheet(ws, header_row)
            sheet = XlsxSheet(name=name, headers=headers, header_row=header_row, rows=rows)
        else:
            sheet = XlsxSheet(name=name, headers=[], header_row=0, rows=[])
        sheet.kv = _parse_kv_sheet(ws)
        sheets[name] = sheet
    return XlsxWorkbook(path=path, filename=path.name, sheets=sheets)


# --- Adapt to schemas for the LLM --------------------------------------------

def _fmt_value(v: Any) -> str:
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if v is None:
        return ""
    return str(v)


def summarize(book: XlsxWorkbook, sample_rows_per_sheet: int = 5) -> XlsxFacts:
    """Return a compact XlsxFacts summary — enough for the judge to reason over."""
    sheets: list[XlsxSheetSummary] = []
    for name, sh in book.sheets.items():
        rows = []
        for row in sh.rows[:sample_rows_per_sheet]:
            rows.append({h: _fmt_value(row.values.get(h)) for h in sh.headers})
        sheets.append(
            XlsxSheetSummary(
                name=name,
                row_count=len(sh.rows),
                col_count=len(sh.headers),
                headers=sh.headers,
                sample_rows=rows,
            )
        )
    return XlsxFacts(file=book.filename, sheets=sheets)


def extract_xlsx_facts(path: Path) -> tuple[XlsxWorkbook, XlsxFacts]:
    """Load + summarize an xlsx workbook."""
    book = load_xlsx(path)
    return book, summarize(book)
